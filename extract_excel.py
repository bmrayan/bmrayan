#!/usr/bin/env python3
import csv
import json

# Try to read the Excel file using different methods
try:
    import pandas as pd
    df = pd.read_excel('FinalDataSet_DGA.xlsx')
    print(f"Using pandas - Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    
    # Convert to JSON format for web
    data = []
    for i, row in df.iterrows():
        record = {
            "ID": i + 1,
            "Hydrogen": float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0.0,
            "Methane": float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0.0,
            "Ethylene": float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0.0,
            "Ethane": float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0.0,
            "Acetylene": float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0.0,
            "Fault": str(row.iloc[6]) if len(row) > 6 and pd.notna(row.iloc[6]) else "NA"
        }
        data.append(record)
    
    # Save as JSON
    with open('excel_data.json', 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f"Extracted {len(data)} records")
    print("Sample record:", data[0] if data else "No data")
    
    # Print fault distribution
    fault_counts = {}
    for record in data:
        fault = record['Fault']
        fault_counts[fault] = fault_counts.get(fault, 0) + 1
    print("Fault distribution:", fault_counts)
    
except ImportError:
    try:
        import openpyxl
        wb = openpyxl.load_workbook('FinalDataSet_DGA.xlsx')
        ws = wb.active
        
        print(f"Using openpyxl - Rows: {ws.max_row}, Cols: {ws.max_column}")
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        print(f"Headers: {headers}")
        
        # Extract data
        data = []
        for row_num in range(2, ws.max_row + 1):
            record = {
                "ID": row_num - 1,
                "Hydrogen": float(ws.cell(row_num, 2).value or 0),
                "Methane": float(ws.cell(row_num, 3).value or 0),
                "Ethylene": float(ws.cell(row_num, 4).value or 0),
                "Ethane": float(ws.cell(row_num, 5).value or 0),
                "Acetylene": float(ws.cell(row_num, 6).value or 0),
                "Fault": str(ws.cell(row_num, 7).value or "NA") if ws.max_column > 6 else "NA"
            }
            data.append(record)
        
        # Save as JSON
        with open('excel_data.json', 'w') as f:
            json.dump(data, f, indent=2)
        
        print(f"Extracted {len(data)} records")
        print("Sample record:", data[0] if data else "No data")
        
        # Print fault distribution
        fault_counts = {}
        for record in data:
            fault = record['Fault']
            fault_counts[fault] = fault_counts.get(fault, 0) + 1
        print("Fault distribution:", fault_counts)
        
    except ImportError:
        print("Neither pandas nor openpyxl available")
        print("Please install one of them: pip install pandas openpyxl")